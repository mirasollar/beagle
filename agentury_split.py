import streamlit as st
import pandas as pd
import openpyxl
import os
from io import BytesIO
import re
from os.path import isfile, join
import zipfile
import shutil

def save_file(uploaded):
    # Určete cestu, kam se soubor uloží
    file_path = os.path.join(os.getcwd(), uploaded.name)
    
    # Získání obsahu souboru jako bytes
    content = uploaded.getvalue()
    
    # Načtení obsahu do pandas DataFrame
    excel_data = pd.read_excel(BytesIO(content), engine='openpyxl')
    
    # Uložení souboru na disk
    with open(file_path, "wb") as f:
        f.write(content)
    return file_path, excel_data

def make_dir(dirn):
    try:
        os.makedirs(dirn)
    except OSError:
        pass

make_dir("split_files")

def get_sheetnames_xlsx(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames

def remove_file(file_name):
    if os.path.isfile(file_name):
        os.remove(file_name)
    else:
        pass

uploaded_file = st.file_uploader("Choose a file")

if uploaded_file is not None and uploaded_file.type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
    fpath=save_file(uploaded_file)
    # st.write(f"Uploaded file name: {uploaded_file.name}")
    uploaded_file_name = uploaded_file.name
    # st.write(f"Sheet name: {get_sheetnames_xlsx(uploaded_file_name)[0]}")

    df_raw = pd.read_excel(uploaded_file_name, nrows=20)
    header_loc = df_raw[df_raw == 'Agentura'].dropna(axis=1, how='all').dropna(how='all')
    header_row = header_loc.index.item()
    df_agentury = pd.read_excel(uploaded_file_name, header=header_row+1)
    # st.write(df_agentury)

    df_agentury_rollup_loc = df_agentury[df_agentury['Agentura'] == 'Rollup']

    df_agentury_start_loc = df_agentury[df_agentury['Agentura'] != '']

    df_agentury_end_loc = df_agentury[df_agentury['Klient'] == 'Rollup']

    names_list = df_agentury_start_loc["Agentura"].to_list()
    names = list(dict.fromkeys(names_list))
    del names[-1]

    del names_list[-1]

    def my_index_multi(l, x):
        return [i for i, _x in enumerate(l) if _x == x]

    start_list = []

    for i in range(len(names)):
        name = names[i]
        idx_list = my_index_multi(names_list, name)
        start_list.append(idx_list[0]+1)

    end_list = []
    for row in df_agentury_end_loc.index:
        end_list.append(row+1)

    source_file = uploaded_file_name
    item_num = len(start_list)

    st.write("Creating split files...")
    for i in range(len(start_list)):
        num_mod = i+1
        if num_mod % 10 == 0:
            st.write(f"Number of finished files: {i+1} of the total {len(start_list)}")
        start_num = start_list[i]
        end_num = end_list[i]+1
        name = names[i]
        name = re.sub('s\.\s?r\.\s?o\.\s?', 'sro', name)
        name = re.sub('a\.\s?s\.\s?', 'as', name)
        name = re.sub('\.|,', '', name)
        name = re.sub('\s', '_', name).lower()
    #     print(name)
        book = openpyxl.load_workbook(source_file)
        sheet = book[get_sheetnames_xlsx(uploaded_file_name)[0]]
        sheet.delete_rows(1, 4)
        
        if i == 0:
    #         print(f"First: {name}, End num: {end_num+1}")
            sheet.delete_rows(end_num+1, 10000)
        elif i+1 == item_num:
            sheet.delete_rows(end_num+1, end_num+1)
            sheet.delete_rows(2, start_num-1)
        else:
            # není první ani poslední
            sheet.delete_rows(end_num+1, 10000)
            sheet.delete_rows(2, start_num-1)
            
        sheet.delete_cols(1, 1)
        sheet.delete_cols(2, 1)

        output_file = f'split_files/{name}.xlsx'
        book.save(output_file)

    cwd = os.getcwd()
    # st.write(f"Current dir: {cwd}")
    result_path = os.path.join(cwd, "split_files")
    # st.write(f"Output path: {result_path}")

    # st.write(f"Obsah output_path: {os.listdir(result_path)}")

    agentury = [f for f in os.listdir(result_path) if isfile(join(result_path, f))]

    zip_file = zipfile.ZipFile('agentury.zip', 'w')
    for i in range(len(agentury)):
        agentura = agentury[i]
        results_agentura = f"split_files/{agentura}"
        # st.write(results_agentura)
        zip_file.write(results_agentura)
    zip_file.close()
    uploaded_file
    # st.write(f"Obsah current dir: {os.listdir()}")

    # uploaded_file = None

with open("agentury.zip", "rb") as fp:
    btn = st.download_button(
            label="Download ZIP",
            data=fp,
            file_name="agentury.zip",
            mime="application/zip"
        )

        try:
            shutil.rmtree('split_files')
        except:
            pass
